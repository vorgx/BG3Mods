#!/usr/bin/env python3
"""
Read Warrior Progression Excel file and output structured data
Applies parsing rule: Extract only text BEFORE the "-" (dash) for ability names
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

def extract_ability_name(cell_value):
    """Extract ability name (text before '-') from cell value"""
    if cell_value is None:
        return None
    
    text = str(cell_value).strip()
    if not text or text == 'None':
        return None
    
    # Split on '-' and take first part
    if '-' in text:
        return text.split('-')[0].strip()
    return text

def read_excel_data(excel_path):
    """Read the Warrior Progression Excel file"""
    
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        return None
    
    print(f"Reading Excel file: {excel_path}\n")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        print(f"Worksheet: {ws.title}")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns\n")
        
        # Read column headers (Row 3 typically has section headers)
        print("=" * 80)
        print("SECTION HEADERS (Row 3):")
        print("=" * 80)
        
        headers = {}
        for col in range(1, min(ws.max_column + 1, 72)):  # Read up to column 72
            cell_value = ws.cell(row=3, column=col).value
            if cell_value:
                headers[col] = str(cell_value).strip()
                print(f"Column {col} ({chr(64+col) if col <= 26 else 'Col'+str(col)}): {headers[col]}")
        
        print("\n" + "=" * 80)
        print("FURY L1 BASELINE (Columns I-J, Row 4 for L1):")
        print("=" * 80)
        
        # Read Fury L1 (Row 4, Columns I=9, J=10)
        fury_l1_col_i = ws.cell(row=4, column=9).value
        fury_l1_col_j = ws.cell(row=4, column=10).value
        
        print(f"Column I (9): {fury_l1_col_i}")
        print(f"  -> Parsed: {extract_ability_name(fury_l1_col_i)}")
        print(f"Column J (10): {fury_l1_col_j}")
        print(f"  -> Parsed: {extract_ability_name(fury_l1_col_j)}")
        
        print("\n" + "=" * 80)
        print("CLASS TALENTS ROW 1 (Columns K-L, Row 4 for L1):")
        print("=" * 80)
        
        # Read Class Talents Row 1 (Row 4, Columns K=11, L=12)
        class_row1_k = ws.cell(row=4, column=11).value
        class_row1_l = ws.cell(row=4, column=12).value
        
        print(f"Column K (11): {class_row1_k}")
        print(f"  -> Parsed: {extract_ability_name(class_row1_k)}")
        print(f"Column L (12): {class_row1_l}")
        print(f"  -> Parsed: {extract_ability_name(class_row1_l)}")
        
        print("\n" + "=" * 80)
        print("COLOSSUS HERO TALENTS (Columns BG-BI):")
        print("=" * 80)
        
        # Read Colossus L13-L20 (Columns BG=59, BH=60, BI=61)
        # Row mapping: Row 4=L1, so L13=Row 16
        colossus_data = {}
        level_map = {
            16: 13,  # Row 16 = L13
            17: 14,  # Row 17 = L14
            18: 15,  # Row 18 = L15
            19: 16,  # Row 19 = L16
            20: 17,  # Row 20 = L17
            21: 18,  # Row 21 = L18
            22: 19,  # Row 22 = L19
            23: 20   # Row 23 = L20
        }
        
        for row, level in level_map.items():
            col_bg = extract_ability_name(ws.cell(row=row, column=59).value)
            col_bh = extract_ability_name(ws.cell(row=row, column=60).value)
            col_bi = extract_ability_name(ws.cell(row=row, column=61).value)
            
            talents = [t for t in [col_bg, col_bh, col_bi] if t]
            if talents:
                colossus_data[level] = talents
                print(f"L{level}: {' + '.join(talents)}")
        
        print("\n" + "=" * 80)
        print("MOUNTAIN THANE HERO TALENTS (Columns BD-BF):")
        print("=" * 80)
        
        # Read Mountain Thane L13-L20 (Columns BD=56, BE=57, BF=58)
        mt_data = {}
        for row, level in level_map.items():
            col_bd = extract_ability_name(ws.cell(row=row, column=56).value)
            col_be = extract_ability_name(ws.cell(row=row, column=57).value)
            col_bf = extract_ability_name(ws.cell(row=row, column=58).value)
            
            talents = [t for t in [col_bd, col_be, col_bf] if t]
            if talents:
                mt_data[level] = talents
                print(f"L{level}: {' + '.join(talents)}")
        
        print("\n" + "=" * 80)
        print("SLAYER HERO TALENTS (Columns BJ-BL):")
        print("=" * 80)
        
        # Read Slayer L13-L20 (Columns BJ=62, BK=63, BL=64)
        slayer_data = {}
        for row, level in level_map.items():
            col_bj = extract_ability_name(ws.cell(row=row, column=62).value)
            col_bk = extract_ability_name(ws.cell(row=row, column=63).value)
            col_bl = extract_ability_name(ws.cell(row=row, column=64).value)
            
            talents = [t for t in [col_bj, col_bk, col_bl] if t]
            if talents:
                slayer_data[level] = talents
                print(f"L{level}: {' + '.join(talents)}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"ERROR reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_path = r"C:\Users\tenod\source\repos\BG3Mods\Documentation\00_SourcesOfTruth\Warrior Progression for all subclasses.xlsx"
    read_excel_data(excel_path)
